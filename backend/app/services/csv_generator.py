"""
Table extraction and export for CSV/XLSX downloads.

Uses BeautifulSoup for reliable colspan/rowspan handling.
This module is EXPORT-ONLY - does not affect RAG embedding pipeline.
"""

from dataclasses import dataclass
from io import StringIO, BytesIO
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
import csv
import re
import logging

logger = logging.getLogger(__name__)


@dataclass
class TableData:
    """Structured representation of a table for CSV/XLSX export."""
    headers: list[str]
    rows: list[list[str]]


def extract_cell_text(cell) -> str:
    """Extract clean text from table cell, handling nested HTML."""
    text = cell.get_text(separator=" ", strip=True)
    # Collapse multiple whitespace
    text = " ".join(text.split())
    return text


def handle_merged_cells(rows) -> list[list[str]]:
    """
    Convert table with colspan/rowspan to regular 2D grid.
    
    IMPORTANT: Only the first cell of a merged region gets content.
    Continuation cells are left empty to avoid duplication.
    """
    if not rows:
        return []
    
    # Determine grid dimensions
    max_cols = 0
    for row in rows:
        cols = sum(int(cell.get('colspan', 1)) for cell in row.find_all(['td', 'th']))
        max_cols = max(max_cols, cols)
    
    if max_cols == 0:
        return []
    
    row_count = len(rows)
    grid = [[None] * max_cols for _ in range(row_count)]
    
    # Fill grid
    for row_idx, row in enumerate(rows):
        col_idx = 0
        for cell in row.find_all(['td', 'th']):
            # Skip already-filled cells (from previous rowspan)
            while col_idx < max_cols and grid[row_idx][col_idx] is not None:
                col_idx += 1
            
            if col_idx >= max_cols:
                break
            
            colspan = int(cell.get('colspan', 1))
            rowspan = int(cell.get('rowspan', 1))
            text = extract_cell_text(cell)
            
            # FIX: Only first cell of merged region gets content
            # Continuation cells stay empty to avoid duplication
            for r in range(min(rowspan, row_count - row_idx)):
                for c in range(min(colspan, max_cols - col_idx)):
                    if r == 0 and c == 0:
                        grid[row_idx + r][col_idx + c] = text  # First cell gets content
                    else:
                        grid[row_idx + r][col_idx + c] = ""    # Others stay empty
            
            col_idx += colspan
    
    # Replace None with empty string
    return [[cell if cell is not None else "" for cell in row] for row in grid]


def collapse_empty_columns(headers: list[str], rows: list[list[str]]) -> tuple[list[str], list[list[str]]]:
    """
    Remove columns that are entirely empty (from merged cell handling).
    
    This cleans up the gaps left by colspan handling without risking data loss.
    Only removes columns where EVERY cell (header + all rows) is empty/whitespace.
    """
    if not headers:
        return headers, rows
    
    num_cols = len(headers)
    
    # Find columns that have at least one non-empty cell
    non_empty_cols = []
    for col_idx in range(num_cols):
        # Check header
        has_content = bool(headers[col_idx].strip())
        
        # Check all rows if header is empty
        if not has_content:
            for row in rows:
                if col_idx < len(row) and row[col_idx].strip():
                    has_content = True
                    break
        
        if has_content:
            non_empty_cols.append(col_idx)
    
    # If all columns have content (or none do), return as-is
    if len(non_empty_cols) == num_cols or len(non_empty_cols) == 0:
        return headers, rows
    
    # Filter to only non-empty columns
    new_headers = [headers[i] for i in non_empty_cols]
    new_rows = [
        [row[i] if i < len(row) else "" for i in non_empty_cols]
        for row in rows
    ]
    
    return new_headers, new_rows


def is_data_table(table_element) -> bool:
    """
    Filter layout tables from data tables using financial heuristics.
    
    Returns True if the table appears to contain actual data (financial tables,
    data grids) rather than being used for layout purposes.
    """
    rows = table_element.find_all('tr')
    cells = table_element.find_all(['td', 'th'])
    text = table_element.get_text(strip=True)
    
    # Must have structure
    if len(rows) < 2 or len(cells) < 6 or len(text) < 50:
        return False
    
    # Financial content indicators
    has_currency = bool(re.search(r'[\$€£¥]', text))
    has_formatted_nums = bool(re.search(r'\d{1,3}(,\d{3})+', text))
    has_percentages = bool(re.search(r'\d+\.?\d*\s*%', text))
    num_count = len(re.findall(r'\d+', text))
    
    return has_currency or has_formatted_nums or has_percentages or num_count >= 8


def extract_table_at_index(html: str, index: int) -> TableData | None:
    """
    Extract a specific table from HTML by index.
    
    Uses BeautifulSoup with fixed colspan/rowspan handling that doesn't
    duplicate content across merged cells.
    
    Args:
        html: Raw HTML content
        index: Zero-based index of the table to extract
        
    Returns:
        TableData if table exists at index and is a data table, None otherwise
    """
    soup = BeautifulSoup(html, "html.parser")
    tables = soup.find_all("table")
    
    if index < 0 or index >= len(tables):
        return None
    
    table_element = tables[index]
    
    # Validate it's a data table, not layout
    if not is_data_table(table_element):
        logger.warning(f"Table {index} appears to be a layout table, skipping export")
        return None
    
    rows = table_element.find_all("tr")
    
    if not rows:
        return None
    
    # Detect headers (first row with <th> OR first row if no <th>)
    header_row_idx = 0
    headers = []
    
    for idx, row in enumerate(rows):
        ths = row.find_all("th")
        if ths:
            header_row_idx = idx
            headers = [extract_cell_text(th) for th in ths]
            break
    
    # If no <th> found, use first row as headers
    if not headers and rows:
        first_row_cells = rows[0].find_all(['td', 'th'])
        if first_row_cells:
            headers = [extract_cell_text(cell) for cell in first_row_cells]
            header_row_idx = 0
        else:
            # No headers at all, generate generic ones
            first_data_cells = rows[1].find_all(['td', 'th']) if len(rows) > 1 else []
            col_count = len(first_data_cells) if first_data_cells else 1
            headers = [f"Column {i+1}" for i in range(col_count)]
    
    # Extract data rows (all rows after header)
    data_rows = rows[header_row_idx + 1:] if len(rows) > header_row_idx + 1 else []
    
    if not data_rows:
        # Table with headers only
        return TableData(headers=headers, rows=[])
    
    # Handle merged cells to create regular grid (with fix for duplication)
    grid = handle_merged_cells(data_rows)
    
    if not grid:
        return TableData(headers=headers, rows=[])
    
    # Remove completely empty rows
    grid = [row for row in grid if any(cell.strip() for cell in row)]
    
    # Collapse empty columns left by merged cell handling
    headers, grid = collapse_empty_columns(headers, grid)
    
    return TableData(headers=headers, rows=grid)


def generate_csv(table: TableData) -> str:
    """
    Generate CSV content from table data.
    
    Args:
        table: Structured table data
        
    Returns:
        CSV-formatted string with proper escaping
    """
    output = StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
    
    # Write headers
    writer.writerow(table.headers)
    
    # Write data rows
    for row in table.rows:
        writer.writerow(row)
    
    return output.getvalue()


def generate_xlsx(table: TableData, source_url: str = None) -> bytes:
    """
    Generate XLSX content from table data with formatting.
    
    Features:
    - Bold header row
    - Optional source link at bottom for traceability
    
    Args:
        table: Structured table data
        source_url: Optional URL to include as "View Source" link
        
    Returns:
        XLSX file content as bytes
    """
    wb = Workbook()
    ws = wb.active
    
    # Write headers with bold formatting
    for col, header in enumerate(table.headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Write data rows
    for row_idx, row in enumerate(table.rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add source link at bottom if provided
    if source_url:
        link_row = len(table.rows) + 3  # 2 blank rows after data
        cell = ws.cell(row=link_row, column=1, value="View Source")
        cell.hyperlink = source_url
        cell.font = Font(color="0563C1", underline="single")
    
    # Write to bytes buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_table_link(filing_url: str, table_index: int, base_url: str = "http://localhost:3000") -> str:
    """
    Build a deep link URL that opens the filing and highlights the specific table.
    
    Args:
        filing_url: Original SEC filing URL
        table_index: Index of the table
        base_url: Base URL of the frontend app
        
    Returns:
        Complete URL with selection parameter
    """
    import base64
    import json
    from urllib.parse import urlencode
    
    selection = {"type": "table", "tableIndex": table_index}
    encoded = base64.b64encode(json.dumps(selection).encode()).decode()
    
    params = urlencode({"source": filing_url, "selection": encoded})
    return f"{base_url}/view?{params}"
